"""Module for processing quiz data from Excel files."""
import pandas as pd
import re
from pathlib import Path
from openpyxl.styles import PatternFill
from typing import List, Dict, Set, Tuple, Optional
from ui.score_change import ScoreChange

class QuizProcessor:
    """Class for processing quiz data from Excel files."""
    
    HIGHLIGHT_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    def __init__(self, input_file: Path, sheet_name: str, total_points: float, raw_score_per_question: float):
        """Initialize the quiz processor with quiz parameters."""
        self.input_file = input_file
        self.sheet_name = sheet_name
        self.total_points = total_points
        self.raw_score_per_question = raw_score_per_question
        self.df = None
        self.question_numbers = []
        self.max_possible_raw_total = 0
        self.changed_scores: Dict[str, Set[int]] = {}
        self._load_data()
    
    def _load_data(self) -> None:
        """Load data from Excel file and extract question numbers."""
        self.df = pd.read_excel(self.input_file, sheet_name=self.sheet_name)
        self.question_numbers = self._extract_question_numbers()
        self.max_possible_raw_total = len(self.question_numbers) * self.raw_score_per_question
    
    def _extract_question_numbers(self) -> List[int]:
        """Extract unique question numbers from column names."""
        question_numbers = set()
        for col in self.df.columns:
            match = re.match(r'(\d+)_score', col.lower())
            if match:
                question_numbers.add(int(match.group(1)))
        return sorted(list(question_numbers))

    def record_score_changes(self, changes: List['ScoreChange']) -> None:
        """Record which scores were changed for highlighting."""
        self.changed_scores = {}
        for change in changes:
            if change.team_name not in self.changed_scores:
                self.changed_scores[change.team_name] = set()
            self.changed_scores[change.team_name].add(change.question_number)
    
    def _get_team_scores(self, team_data) -> Tuple[List[float], float, float]:
        """Calculate team scores from the first student's data."""
        first_student = team_data.iloc[0]
        earned_raw_scores = []
        
        for q_num in self.question_numbers:
            score_col = f"{q_num}_Score"
            if score_col in self.df.columns:
                earned_raw_score = float(first_student[score_col])
                earned_raw_scores.append(earned_raw_score)
        
        team_raw_total = sum(earned_raw_scores)
        team_adjusted_total = (team_raw_total * self.total_points) / self.max_possible_raw_total
        
        return earned_raw_scores, team_raw_total, team_adjusted_total
    
    def _calculate_adjusted_scores(self, earned_raw_scores: List[float]) -> List[float]:
        """Calculate adjusted scores for individual questions."""
        points_per_question = self.total_points / len(self.question_numbers)
        return [
            (earned_score * points_per_question) / self.raw_score_per_question 
            for earned_score in earned_raw_scores
        ]
    
    def process_data(self) -> Tuple[List[Dict], List[int], float]:
        """Process quiz data and calculate scores."""
        results = []
        
        for team_name, team_data in self.df.groupby('Team'):
            earned_raw_scores, team_raw_total, team_adjusted_total = self._get_team_scores(team_data)
            adjusted_scores = self._calculate_adjusted_scores(earned_raw_scores)
            
            for _, student in team_data.iterrows():
                results.append({
                    'Team Name': team_name,
                    'Student ID': student['Student ID'],
                    'Student Name': student['Student Name'],
                    'Email Address': student['Email Address'],
                    'Raw Scores': earned_raw_scores,
                    'Student Raw Total': team_raw_total,
                    'Team Raw Total': team_raw_total,
                    'Team Adjusted Total': team_adjusted_total,
                    'Adjusted Scores': adjusted_scores,
                    'Student Adjusted Total': team_adjusted_total
                })
        
        return results, self.question_numbers, self.max_possible_raw_total

    def _create_output_dataframe(self, results: List[Dict]) -> pd.DataFrame:
        """Create DataFrame for output."""
        output_data = []
        current_team = None
        
        for result in results:
            is_first_in_team = result['Team Name'] != current_team
            
            row = {
                'Team Name': result['Team Name'] if is_first_in_team else '',
                'Team Raw Total': result['Team Raw Total'] if is_first_in_team else '',
                'Team Adjusted Total': result['Team Adjusted Total'] if is_first_in_team else '',
                'Student ID': result['Student ID'],
                'Student Name': result['Student Name'],
                'Email Address': result['Email Address'],
                'Student Raw Total': result['Student Raw Total'],
                'Student Adjusted Total': result['Student Adjusted Total']
            }
            current_team = result['Team Name']
            
            for i, (raw, adjusted) in enumerate(zip(result['Raw Scores'], result['Adjusted Scores']), 1):
                row[f'Q{i} Raw Score'] = raw
                row[f'Q{i} Adjusted Score'] = adjusted
            
            output_data.append(row)
        
        return pd.DataFrame(output_data)

    def _get_score_column_indices(self, df: pd.DataFrame) -> Tuple[Dict[int, int], Dict[int, int]]:
        """Get column indices for raw and adjusted score columns."""
        raw_score_cols = {
            q: next(i for i, col in enumerate(df.columns, 1) 
                   if col == f'Q{q} Raw Score')
            for q in self.question_numbers
        }
        adjusted_score_cols = {
            q: next(i for i, col in enumerate(df.columns, 1) 
                   if col == f'Q{q} Adjusted Score')
            for q in self.question_numbers
        }
        return raw_score_cols, adjusted_score_cols    
    
    def _highlight_changed_scores(self, worksheet, df: pd.DataFrame) -> None:
        """Apply highlighting to changed scores in worksheet."""
        if not self.changed_scores:
            return
            
        raw_score_cols, adjusted_score_cols = self._get_score_column_indices(df)
        
        # Track current team for empty team name rows
        current_team = None
        for row_idx, row in enumerate(df.to_dict('records'), 2):
            # Update current team when we find a non-empty team name
            if row['Team Name']:
                current_team = row['Team Name']
            
            # Use current team name for highlighting all members
            if current_team in self.changed_scores:
                for q_num in self.changed_scores[current_team]:
                    for col_idx in [raw_score_cols[q_num], adjusted_score_cols[q_num]]:
                        worksheet.cell(row=row_idx, column=col_idx).fill = self.HIGHLIGHT_FILL

    def save_to_excel(self, results: List[Dict], output_file: Path) -> None:
        """Save processed data to Excel file with highlighting."""
        df_output = self._create_output_dataframe(results)
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_output.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            self._highlight_changed_scores(worksheet, df_output)

    @classmethod
    def create_output_excel(cls, results: List[Dict], question_numbers: List[int], 
                          output_file: Path, processor: Optional['QuizProcessor'] = None) -> None:
        """Create the output Excel file with the processed data."""
        if processor:
            processor.save_to_excel(results, output_file)
        else:
            df_output = cls._create_output_dataframe(results)
            df_output.to_excel(output_file, index=False)